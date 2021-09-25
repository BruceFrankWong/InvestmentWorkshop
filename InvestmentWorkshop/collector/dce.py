# -*- coding: UTF-8 -*-

__author__ = 'Bruce Frank Wong'


from typing import Dict, List, Any
import datetime as dt
from pathlib import Path
import csv

import requests
from lxml import etree
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import coordinate_from_string

from ..utility import CONFIGS
from .utility import QuoteDaily, split_symbol


DCE_History_URL_Index = Dict[int, Dict[str, str]]


def fetch_dce_history_index() -> DCE_History_URL_Index:
    """
    Download history data (monthly) from DCE.
    :return: Dict[int, Dict[str, str]].
    """
    result: DCE_History_URL_Index = {}
    url_dce: str = 'http://www.dce.com.cn'
    url: str = f'{url_dce}/dalianshangpin/xqsj/lssj/index.html'

    # Make sure <year> in possible range.
    year_list: List[int] = [year for year in range(2006, dt.date.today().year)]
    year_list.reverse()
    for year in year_list:
        result[year] = {}

    # Download index page.
    response = requests.get(url)
    if response.status_code != 200:
        raise requests.exceptions.HTTPError(
            f'Something wrong in downloading <{url}>.'
        )
    response.encoding = 'utf-8'

    # Parse.
    html: etree._Element = etree.HTML(response.text)
    history_data_list: List[etree._Element] = html.xpath('//ul[@class="cate_sel clearfix"]')

    for i in range(len(year_list)):
        product_list = history_data_list[i].xpath('./li/label/text()')
        url_list = history_data_list[i].xpath('./li/label/input/@rel')
        assert len(product_list) == len(url_list)
        for j in range(len(product_list)):
            result[year_list[i]][product_list[j]] = f'{url_dce}/{url_list[j]}'

    return result


def download_dce_history_data(year: int) -> None:
    data_index: DCE_History_URL_Index = fetch_dce_history_index()
    download_path: Path = Path(CONFIGS['path']['download'])
    extension_name: str
    for product, url in data_index[year].items():
        extension_name = url.split('.')[-1]
        download_file = download_path.joinpath(f'DCE_{product}_{year}.{extension_name}')
        response = requests.get(url)
        if response.status_code != 200:
            raise requests.exceptions.HTTPError(
                f'Something wrong in downloading <{url}>.'
            )
        with open(download_file, 'wb') as f:
            f.write(response.content)


def download_dce_history_data_all() -> None:
    start_year: int = 2006
    this_year: int = dt.date.today().year
    for year in range(start_year, this_year):
        download_dce_history_data(year)


def correct_format(file_path: Path) -> str:
    """
    Return the correct file format.
    :param file_path: a Path-like object.
    :return: str.
    """
    if file_path.suffix == '.csv':
        with open(file=file_path, mode='rb') as csv_file:
            header = csv_file.read(4)
        # .xlsx
        if header == b'\x50\x4B\x03\x04':
            return '.xlsx'
        # .xls
        if header == b'\x3C\x3F\x78\x6D':
            return '.xls'
    return '.csv'


def read_dce_history_data_xls(xls_path: Path) -> List[QuoteDaily]:
    """
    Read quote data from .xls file.
    :param xls_path: a Path-like object.
    :return: List[QUOTE].
    """
    assert xls_path.exists() is True

    # It seemed that only
    assert '期权' in xls_path.stem

    result: List[QuoteDaily] = []

    # Read .xls files.
    workbook = xlrd.open_workbook(xls_path)
    for data_sheet in workbook.sheets():

        # Columns.
        xls_column_list: List[str] = [x.value for x in data_sheet.row(0)]
        for i in range(1, data_sheet.nrows):
            row = data_sheet.row(i)
            date_str = row[xls_column_list.index('交易日期')].value
            if date_str is None:
                print(f'Error in {xls_path} at row {i}')
            result.append(
                {
                    'exchange': 'DCE',
                    'symbol': row[xls_column_list.index('合约名称')].value,
                    'date': dt.date(
                        year=int(date_str[:4]),
                        month=int(date_str[4:6]),
                        day=int(date_str[6:8])
                    ),
                    'open': float(row[xls_column_list.index('开盘价')].value),
                    'high': float(row[xls_column_list.index('最高价')].value),
                    'low': float(row[xls_column_list.index('最低价')].value),
                    'close': float(row[xls_column_list.index('收盘价')].value),
                    'previous_settlement': float(row[xls_column_list.index('前结算价')].value),
                    'settlement': float(row[xls_column_list.index('结算价')].value),
                    'change_on_close': float(row[xls_column_list.index('涨跌')].value),
                    'change_on_settlement': float(row[xls_column_list.index('涨跌1')].value),
                    'delta': float(row[xls_column_list.index('DELTA')].value)
                    if row[xls_column_list.index('DELTA')].value != '' else None,
                    'volume': int(row[xls_column_list.index('成交量')].value),
                    'open_interest': int(row[xls_column_list.index('持仓量')].value),
                    'change_on_open_interest': int(row[xls_column_list.index('持仓量变化')].value),
                    'amount': float(row[xls_column_list.index('成交额（万元）')].value) * 10000,
                    'exercise': int(row[xls_column_list.index('行权量')].value),
                }
            )

    return result


def read_dce_history_data_xlsx(xlsx_path: Path) -> List[QuoteDaily]:
    """
    Read quote data from .xlsx file.
    :param xlsx_path: a Path-like object.
    :return: List[QUOTE].
    """
    result: List[QuoteDaily] = []
    assert xlsx_path.exists() is True
    workbook: Workbook = load_workbook(filename=xlsx_path)
    worksheet: Worksheet = workbook.active
    column_max: str
    row_max: int
    column_max, row_max = coordinate_from_string(worksheet.dimensions.split(':')[1])
    column_list: List[str] = [str(x.value).strip() for x in list(worksheet.rows)[0]]
    for row in range(2, row_max):
        if '交易日期' in column_list:
            day = worksheet.cell(row=row, column=column_list.index('交易日期') + 1).value
        else:
            day = worksheet.cell(row=row, column=column_list.index('日期') + 1).value
        if day is None:
            break
        else:
            day = dt.date(
                year=int(str(day)[:4]),
                month=int(str(day)[4:6]),
                day=int(str(day)[6:8]),
            )

        symbol: str
        if '合约名称' in column_list:
            symbol = str(worksheet.cell(row=row, column=column_list.index('合约名称') + 1).value)
        else:
            symbol = str(worksheet.cell(row=row, column=column_list.index('合约') + 1).value)

        price_open: Any = worksheet.cell(row=row, column=column_list.index('开盘价') + 1).value
        price_high: Any = worksheet.cell(row=row, column=column_list.index('最高价') + 1).value
        price_low: Any = worksheet.cell(row=row, column=column_list.index('最低价') + 1).value
        price_close: float = float(worksheet.cell(row=row, column=column_list.index('收盘价')+1).value)
        volume: int = int(float(worksheet.cell(row=row, column=column_list.index('成交量') + 1).value))

        if '期权' in xlsx_path.stem:
            result.append(
                {
                    'exchange': 'DCE',
                    'symbol': symbol,
                    'date': day,
                    'open': None if price_open is None or price_open == 0 or price_open == 0.0 else float(price_open),
                    'high': None if price_high is None or price_high == 0 or price_high == 0.0 else float(price_high),
                    'low': None if price_low is None or price_low == 0 or price_low == 0.0 else float(price_low),
                    'close': price_close,
                    'previous_settlement': float(worksheet.cell(row=row, column=column_list.index('前结算价')+1).value),
                    'settlement': float(worksheet.cell(row=row, column=column_list.index('结算价')+1).value),
                    'change_on_close': float(worksheet.cell(row=row, column=column_list.index('涨跌')+1).value),
                    'change_on_settlement': float(worksheet.cell(row=row, column=column_list.index('涨跌1')+1).value),
                    'delta': None
                    if worksheet.cell(row=row, column=column_list.index('DELTA')+1).value is None or
                    len(str(worksheet.cell(row=row, column=column_list.index('DELTA') + 1).value)) == 0
                    else float(worksheet.cell(row=row, column=column_list.index('DELTA')+1).value),
                    'volume': volume,
                    'open_interest': int(worksheet.cell(row=row, column=column_list.index('持仓量')+1).value),
                    'change_on_open_interest': int(
                        worksheet.cell(row=row, column=column_list.index('持仓量变化')+1).value
                    ),
                    'amount': float(worksheet.cell(row=row, column=column_list.index('成交额（万元）')+1).value) * 10000,
                    'exercise': int(float(worksheet.cell(row=row, column=column_list.index('行权量')+1).value)),
                }
            )
        else:
            if '成交金额' in column_list:
                column_list[column_list.index('成交金额')] = '成交额'
            product, contract = split_symbol(symbol)
            result.append(
                {
                    'exchange': 'DCE',
                    'symbol': symbol,
                    'product': product,
                    'contract': contract,
                    'date': day,
                    'open': None if price_open is None or price_open == 0 or price_open == 0.0 else float(price_open),
                    'high': None if price_high is None or price_high == 0 or price_high == 0.0 else float(price_high),
                    'low': None if price_low is None or price_low == 0 or price_low == 0.0 else float(price_low),
                    'close': price_close,

                    'previous_close': float(worksheet.cell(row=row, column=column_list.index('前收盘价')+1).value),
                    'previous_settlement': float(worksheet.cell(row=row, column=column_list.index('前结算价')+1).value),
                    'settlement': float(worksheet.cell(row=row, column=column_list.index('结算价')+1).value),
                    'change_on_close': float(worksheet.cell(row=row, column=column_list.index('涨跌1')+1).value),
                    'change_on_settlement': float(worksheet.cell(row=row, column=column_list.index('涨跌2')+1).value),
                    'volume': int(float(worksheet.cell(row=row, column=column_list.index('成交量')+1).value)),
                    'amount': float(worksheet.cell(row=row, column=column_list.index('成交额')+1).value),
                    'open_interest': int(float(worksheet.cell(row=row, column=column_list.index('持仓量')+1).value)),
                }
            )

    return result


def read_dce_history_data_csv(csv_path: Path) -> List[QuoteDaily]:
    assert csv_path.exists() is True
    result: List[QuoteDaily] = []
    with open(csv_path, mode='r', encoding='gbk') as csv_file:
        reader = csv.DictReader(csv_file)
        if '期权' in csv_path.stem:
            for row in reader:
                result.append(
                    {
                        'exchange': 'DCE',
                        'symbol': row['合约名称'].strip(),
                        'date': dt.date(
                            year=int(row['交易日期'][:4]),
                            month=int(row['交易日期'][4:6]),
                            day=int(row['交易日期'][6:8])
                        ),
                        'open': None if row['开盘价'] == '0' or row['开盘价'] == 0 else float(row['开盘价']),
                        'high': None if row['最高价'] == '0' else float(row['最高价']),
                        'low': None if row['最低价'] == '0' else float(row['最低价']),
                        'close': float(row['收盘价']),
                        'previous_settlement': float(row['前结算价']),
                        'settlement': float(row['结算价']),
                        'change_on_close': float(row['涨跌']),
                        'change_on_settlement': float(row['涨跌1']),
                        'delta': None if row['Delta'] == '' else float(row['Delta']),
                        'volume': int(row['成交量']),
                        'open_interest': int(float(row['持仓量'])),
                        'change_on_open_interest': int(float(row['持仓量变化'])),
                        'amount': float(row['成交额（万元）']) * 10000,
                        'exercise': int(row['行权量']),
                    }
                )
        else:
            for row in reader:
                symbol = row['合约'].strip()
                product, contract = split_symbol(symbol)
                result.append(
                    {
                        'exchange': 'DCE',
                        'symbol': symbol,
                        'product': product,
                        'contract': contract,
                        'date': dt.date(
                            year=int(row['日期'][:4]),
                            month=int(row['日期'][4:6]),
                            day=int(row['日期'][6:8])
                        ),
                        'previous_close': None
                        if row['前收盘价'] == '0' or row['前收盘价'] == 0 or row['前收盘价'] == '' or row['前收盘价'] is None
                        else float(row['前收盘价']),
                        'previous_settlement': float(row['前结算价']),
                        'open': None if row['开盘价'] == '0' or row['开盘价'] == 0 else float(row['开盘价']),
                        'high': None if row['最高价'] == '0' else float(row['最高价']),
                        'low': None if row['最低价'] == '0' else float(row['最低价']),
                        'close': float(row['收盘价']),
                        'settlement': float(row['结算价']),
                        'change_on_close': float(row['涨跌1']),
                        'change_on_settlement': float(row['涨跌2']),
                        'volume': int(row['成交量']),
                        'amount': float(row['成交金额']) * 10000,
                        'open_interest': int(float(row['持仓量'])),
                    }
                )
    return result


def read_dce_history_data(data_file: Path) -> List[QuoteDaily]:
    assert data_file.exists() is True
    extension: str = data_file.suffix
    if extension == '.csv':
        if correct_format(data_file) == '.csv':
            return read_dce_history_data_csv(data_file)
        elif correct_format(data_file) == '.xls':
            return read_dce_history_data_xls(data_file)
        elif correct_format(data_file) == '.xlsx':
            return read_dce_history_data_xlsx(data_file)
        else:
            raise RuntimeError(f'Unknown file type. {data_file}')
    elif extension == '.xls':
        return read_dce_history_data_xls(data_file)
    elif extension == '.xlsx':
        return read_dce_history_data_xlsx(data_file)
    else:
        raise RuntimeError(f'Unknown file type. {data_file}')
